import sys
import json
import os
import PyQt6.QtGui
from PyQt6 import QtWidgets, QtGui, QtCore
from MainWindow import Ui_MainWindow
from create_alley import Ui_create_alley
from PyQt6.QtWidgets import QTableWidgetItem, QMessageBox, QFileDialog
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting import Rule
from re import fullmatch


class AlignDelegate(QtWidgets.QStyledItemDelegate):  # Делегат центрирования по ячейкам
    def initStyleOption(self, option, index):
        super(AlignDelegate, self).initStyleOption(option, index)
        option.displayAlignment = QtCore.Qt.AlignmentFlag.AlignCenter


class ColorDelegate(QtWidgets.QStyledItemDelegate):  # Делегат цвета балки
    def initStyleOption(self, option, index):
        super(ColorDelegate, self).initStyleOption(option, index)
        option.backgroundBrush = QtGui.QColor('red')


class WhiteColorDelegate(QtWidgets.QStyledItemDelegate):  # Делегат цвета балки
    def initStyleOption(self, option, index):
        super(WhiteColorDelegate, self).initStyleOption(option, index)
        option.backgroundBrush = QtGui.QColor('white')


def resize_columns(worksheet):
    if worksheet is None:
        return

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.3
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width


def export_conditional_format(sheet):
    """Условное форматирование"""
    bg_green = PatternFill(fill_type='solid', bgColor="00FF00")
    dxf = DifferentialStyle(fill=bg_green)
    rule = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    rule.formula = ['$D1="ОК"']
    sheet.conditional_formatting.add("A1:D1000", rule)


class TopConst(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, *args, **kwargs):
        super(TopConst, self).__init__(*args, **kwargs)
        self.setupUi(self)
        self.left_list = self.left_list
        self.new_alley = None
        self.topology = None
        self.topology_file = None
        self.current_alley = None
        self.ns_main = None
        self.ns_header = None
        self.buffer = {'alley': {}}
        self.check_box = QtWidgets.QCheckBox()
        self.copy_action = PyQt6.QtGui.QAction("Copy")
        self.paste_action = PyQt6.QtGui.QAction("Paste")
        self.add_filter_action = PyQt6.QtGui.QAction("Добавить фильтр")
        self.delete_filter_action = PyQt6.QtGui.QAction("Удалить фильтр")

        self.init_settings()

    def init_settings(self):
        # Настройки меню
        self.open.setShortcut('Ctrl+O')
        self.create_topology.setShortcut('Ctrl+N')

        # Настройки списка
        self.left_list.setSortingEnabled(True)

        # Настройки основной таблицы
        table = self.tableWidget
        table.setFocusPolicy(QtCore.Qt.FocusPolicy.NoFocus)  # Убирает выделение элемента при нажатии
        delegate = AlignDelegate(table)  # Выравнивание
        table.setItemDelegate(delegate)  # таблицы по центру
        self.upper_table.setItemDelegate(delegate)
        self.unique_filter_table.setItemDelegate(delegate)
        self.extra_filter_table.setItemDelegate(delegate)

        # Настройка верхней левой таблицы
        up_table = self.upper_table
        value = 40
        header = ((up_table.size().width() - (5 * value)) // 3)

        up_table.setColumnWidth(0, header)
        up_table.setColumnWidth(1, value)
        up_table.setColumnWidth(2, header)
        up_table.setColumnWidth(3, value)
        up_table.setColumnWidth(4, header)
        up_table.setColumnWidth(5, value)
        up_table.setColumnWidth(6, value)
        up_table.setColumnWidth(7, value - 2)

        up_table.setSpan(2, 1, 1, 5)
        up_table.setSpan(1, 6, 2, 2)
        up_table.setSpan(0, 6, 1, 2)

        self.check_box.setGeometry(200, 150, 100, 80)
        self.check_box.setStyleSheet("QCheckBox::indicator"
                                     "{"
                                     "width: 20px;"
                                     "height: 20px;"
                                     "margin-left: 60%;"
                                     "}")
        up_table.setCellWidget(1, 6, self.check_box)

        # Настройка таблицы с фильтром для уникальных штрихкодов
        self.unique_filter_table.setSpan(0, 0, 1, 2)
        filter_table_width = 80
        self.unique_filter_table.setColumnWidth(0, filter_table_width)
        self.unique_filter_table.setColumnWidth(1, filter_table_width)

        # Настройка таблицы с фильтром для неуникальных значений
        self.extra_filter_table.setSpan(0, 0, 1, 2)
        self.extra_filter_table.setColumnWidth(0, filter_table_width)
        self.extra_filter_table.setColumnWidth(1, filter_table_width)

        # Привязка функций
        self.delete_button.clicked.connect(self.delete_alley)
        self.left_list.itemClicked.connect(self.item_clicked)
        self.left_list.itemChanged.connect(self.item_changed)
        self.create_button.clicked.connect(self.create_alley_window)
        self.open.triggered.connect(self.menu_open)
        self.alley_change_button.clicked.connect(self.alley_change)
        self.check_box.toggled.connect(self.photo_checkbox_change)
        self.alley_table_export.triggered.connect(self.alley_export)

        # Настройка контекстного меню левого списка
        self.copy_action.setShortcut('Ctrl+C')
        self.paste_action.setShortcut('Ctrl+V')
        self.copy_action.triggered.connect(self.copy_alley)
        self.paste_action.triggered.connect(self.paste_alley)
        self.left_list.addAction(self.copy_action)
        self.left_list.addAction(self.paste_action)

        # Настройка контекстного меню для уникального фильтра
        self.unique_filter_table.addAction(self.add_filter_action)
        self.unique_filter_table.addAction(self.delete_filter_action)

    def write_line(self, sheet, alley, count):
        current_row = sheet.max_row + 1
        sheet.cell(row=current_row, column=1).value = alley
        sheet.cell(row=current_row, column=2).value = count

        sheet.cell(row=current_row, column=1).style = self.ns_main
        sheet.cell(row=current_row, column=2).style = self.ns_main
        sheet.cell(row=current_row, column=3).style = self.ns_main
        sheet.cell(row=current_row, column=4).style = self.ns_main

    def excel_styles(self):
        self.ns_header = NamedStyle(name='header')
        self.ns_header.font = Font(bold=True, size=14)
        thick_border = Side(style='thick', color='000000')
        self.ns_header.border = Border(left=thick_border, top=thick_border, right=thick_border, bottom=thick_border)
        self.ns_header.alignment.horizontal = 'center'

        self.ns_main = NamedStyle(name='main')
        self.ns_main.font = Font(size=12)
        thin_border = Side(style='thin', color='000000')
        self.ns_main.border = Border(left=thin_border, top=thin_border, right=thin_border, bottom=thin_border)
        self.ns_main.alignment.horizontal = 'center'

    def export_header(self, sheet):
        sheet.cell(row=1, column=1).value = "Аллея"
        sheet.cell(row=1, column=2).value = "Количество ячеек"
        sheet.cell(row=1, column=3).value = "Фамилия"
        sheet.cell(row=1, column=4).value = "Статус"

        sheet.cell(row=1, column=1).style = self.ns_header
        sheet.cell(row=1, column=2).style = self.ns_header
        sheet.cell(row=1, column=3).style = self.ns_header
        sheet.cell(row=1, column=4).style = self.ns_header

    def export_final_numbers(self, sheet):
        sheet.cell(row=1, column=7).value = "Всего ячеек"
        sheet.cell(row=1, column=8).value = "=SUM(B2:B1000)"
        sheet.cell(row=4, column=7).value = "Отлётано"
        sheet.cell(row=4, column=8).value = '=СУММЕСЛИ(D2:D1000,"ОК",B2:B1000)'
        sheet.cell(row=4, column=9).value = '=ROUND(H4/H1*100, 2)&"%"'

        sheet.cell(row=1, column=7).style = self.ns_header
        sheet.cell(row=1, column=8).style = self.ns_header
        sheet.cell(row=4, column=7).style = self.ns_header
        sheet.cell(row=4, column=8).style = self.ns_header
        sheet.cell(row=4, column=9).style = self.ns_header

    def export_print_alleys(self, sheet):
        export_list = list()
        for i, alley in enumerate(self.topology['alley']):
            if fullmatch(r'.+_.+', alley):
                current_alley = alley.split('_')[0]
                current_count = 0
                if current_alley not in export_list:
                    for temp_alley in self.topology['alley']:
                        if temp_alley.split('_')[0] == current_alley:
                            current_count += self.topology['alley'][temp_alley]['rows'] * \
                                             self.topology['alley'][temp_alley]['columns']
                    export_list.append(current_alley)
                    self.write_line(sheet, current_alley, current_count)
            else:
                current_alley = alley
                current_count = self.topology['alley'][alley]['rows'] * \
                    self.topology['alley'][alley]['columns']
                self.write_line(sheet, current_alley, current_count)

    def alley_export(self):
        if self.topology:
            export_table = Workbook()
            export_sheet = export_table.active
            export_sheet.title = "Export"

            self.excel_styles()
            export_table.add_named_style(self.ns_header)
            export_table.add_named_style(self.ns_main)

            self.export_header(export_sheet)
            self.export_print_alleys(export_sheet)
            self.export_final_numbers(export_sheet)
            export_conditional_format(export_sheet)

            resize_columns(export_sheet)
            export_sheet.column_dimensions['H'].width = 20
            export_sheet.column_dimensions['I'].width = 20

            """Сохраняем итоговую таблицу"""
            try:
                export_table.save("Export.xlsx")
                button = QMessageBox.information(self, "Отлично!", "Отчёт создан!", QMessageBox.StandardButton.Ok,
                                                 QMessageBox.StandardButton.Open)
                if button == QMessageBox.StandardButton.Open:
                    os.system("Export.xlsx")
                    return
            except PermissionError:
                QMessageBox.warning(self, "Хьюстон, у нас проблемы!", "Файл с таким названием уже используется")
        else:
            QMessageBox.information(self, "Хьюстон, у нас проблемы!", "Стеллаж не выбран!")

    def copy_alley(self):
        if self.left_list.selectedItems():
            self.buffer = {'alley': {}}
            for alley in self.left_list.selectedItems():
                alley_name = alley.text()
                self.buffer['alley'][alley_name] = dict(self.topology['alley'][alley_name])

    def paste_alley(self):
        if self.buffer:
            self.left_list.itemChanged.disconnect(self.item_changed)
            items = [self.left_list.item(x).text() for x in range(self.left_list.count())]

            for alley_name in self.buffer['alley'].keys():
                self.topology['alley'][alley_name + '_copy'] = dict(self.buffer['alley'][alley_name])

                if alley_name + '_copy' not in items:
                    self.left_list.addItem(alley_name + '_copy')
                    item = self.left_list.item(self.left_list.count() - 1)
                    item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                    item.setFlags(item.flags() | QtCore.Qt.ItemFlag.ItemIsEditable)

            with open(self.topology_file, 'w') as fd:
                json.dump(self.topology, fd)
            fd.close()
            self.buffer = None
            self.left_list.itemChanged.connect(self.item_changed)

    def menu_open(self):  # Кнопка "Открыть" в меню
        self.topology_file = None
        self.topology_file = QFileDialog.getOpenFileName(self, "Open Topology", "", "Text Files (*.txt)")[0]
        if self.topology_file:
            self.topology = self.open_file(str(self.topology_file))
        else:
            return

    def add_beam(self, col):  # Добавляет одну балку
        table = self.tableWidget

        table.insertColumn(col)
        table.horizontalHeader().resizeSection(col, 10)
        delegate = ColorDelegate(table)
        table.setItemDelegateForColumn(col, delegate)

    def add_beams(self, beam_list, local_dir):  # Добавляет ВСЕ балки
        count = 1
        self.add_beam(count)

        if local_dir == 0:
            for n in beam_list:
                count += n + 1
                self.add_beam(count)
        else:
            for n in beam_list[::-1]:
                count += n + 1
                self.add_beam(count)

    def table_headers(self, col_count, row_count, start_level, start_cell, local_direction):
        table = self.tableWidget

        # Горизонтальные заголовки
        if local_direction == 0:
            num = start_cell - col_count + 2
            for col in range(1, col_count):
                table.setItem(0, col, QTableWidgetItem(str(num)))
                num += 1
        elif local_direction == 1:
            num = col_count + start_cell - 2
            for col in range(1, col_count):
                table.setItem(0, col, QTableWidgetItem(str(num)))
                num -= 1

        # Вертикальные заголовки
        row_num = start_level
        for row in range(row_count - 1, 0, -1):
            table.setItem(row, 0, QTableWidgetItem(str(row_num)))
            row_num += 1

    def create_table(self, row_count, col_count, beam_list, start_level, start_cell, local_direction):
        table = self.tableWidget

        # Очищаем таблицу от старых значений
        self.clear_table()
        row_count += 1
        col_count += 1

        # Параметры таблицы
        table.setRowCount(row_count)
        table.setColumnCount(col_count)

        self.table_headers(col_count, row_count, start_level, start_cell, local_direction)
        self.add_beams(beam_list, local_direction)

    def open_file(self, file):
        table = self.upper_table
        self.clear_table()
        self.left_list.clear()

        TopConst.setWindowTitle(self, file)  # В названии окна будет местоположение файла с топологией
        with open(file) as content:
            topology = json.load(content)
        # Вырубаем обработчик событий от греха подальше, точнее от ложных срабатываний
        self.left_list.itemChanged.disconnect(self.item_changed)

        for key in topology['alley']:  # Закидываем аллеи в список слева
            self.left_list.addItem(key)
        self.left_list.sortItems(QtCore.Qt.SortOrder.AscendingOrder)
        for i in range(self.left_list.count()):
            item = self.left_list.item(i)
            item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsSelectable |
                          QtCore.Qt.ItemFlag.ItemIsEditable |
                          QtCore.Qt.ItemFlag.ItemIsEnabled)
        # Включаем обработчик событий обратно
        self.left_list.itemChanged.connect(self.item_changed)

        # Очистка верхней таблицы
        table.setItem(0, 1, QTableWidgetItem(""))
        table.setItem(0, 3, QTableWidgetItem(""))
        table.setItem(1, 1, QTableWidgetItem(""))
        table.setItem(1, 3, QTableWidgetItem(""))
        table.setItem(1, 5, QTableWidgetItem(""))
        table.setItem(2, 1, QTableWidgetItem(""))

        # Очистка и настройка таблиц с фильтрами:
        self.unique_filter_table.setRowCount(1)
        self.unique_filter_table.setRowCount(len(topology['uniq_bar']) + 1)
        self.extra_filter_table.setRowCount(1)
        self.extra_filter_table.setRowCount(len(topology['extra_bar']) + 1)
        for i in range(0, len(topology['uniq_bar'])):
            self.unique_filter_table.setItem(i + 1, 0, QTableWidgetItem(str(topology['uniq_bar'][i][0])))
            self.unique_filter_table.setItem(i + 1, 1, QTableWidgetItem(str(topology['uniq_bar'][i][1])))
        for i in range(0, len(topology['extra_bar'])):
            self.extra_filter_table.setItem(i + 1, 0, QTableWidgetItem(str(topology['extra_bar'][i][0])))
            self.extra_filter_table.setItem(i + 1, 1, QTableWidgetItem(str(topology['extra_bar'][i][1])))

        if topology['photo'] == 2:
            self.check_box.setCheckState(QtCore.Qt.CheckState.Checked)
        elif topology['photo'] == 0:
            self.check_box.setCheckState(QtCore.Qt.CheckState.Unchecked)
        else:
            self.check_box.setCheckState(QtCore.Qt.CheckState.Unchecked)

        return topology

    def item_clicked(self):
        table = self.upper_table
        item = self.left_list.currentItem()
        self.current_alley = item.text()
        alley = self.topology['alley'][item.text()]
        self.create_table(alley['rows'], alley['columns'], alley['list_of_balks'], alley['start_level'],
                          alley['start_cell'], alley['local_direction'])

        table.setItem(0, 1, QTableWidgetItem(str(alley['rows'])))
        table.setItem(0, 3, QTableWidgetItem(str(alley['columns'])))
        table.setItem(0, 5, QTableWidgetItem(str(alley['start_level'])))
        table.setItem(1, 1, QTableWidgetItem(str(alley['start_cell'])))
        table.setItem(1, 3, QTableWidgetItem(str(alley['count_of_barcodes'])))
        table.setItem(1, 5, QTableWidgetItem(str(alley['local_direction'])))
        table.setItem(2, 1, QTableWidgetItem(str(alley['list_of_balks'])))

    def item_changed(self):
        print('CHANGED')
        if self.left_list.currentItem():
            alley = self.topology['alley'][self.current_alley]
            self.topology["alley"].pop(self.current_alley)
            self.topology['alley'][self.left_list.currentItem().text()] = alley
            with open(self.topology_file, 'w') as fd:
                json.dump(self.topology, fd)
            fd.close()
            self.left_list.sortItems(order=QtCore.Qt.SortOrder.AscendingOrder)

    def clear_table(self):
        table = self.tableWidget
        delegate = WhiteColorDelegate(table)
        table.setItemDelegate(delegate)
        for col in range(table.columnCount()):
            table.setItemDelegateForColumn(col, delegate)
        table.resizeColumnsToContents()
        table.setRowCount(0)
        table.setColumnCount(0)

    def delete_alley(self):
        left_list = self.left_list
        if left_list.selectedItems():
            for item in left_list.selectedItems():
                alley = item.text()
                row = left_list.row(item)
                self.topology["alley"].pop(alley)
                left_list.takeItem(row)
            with open(self.topology_file, 'w') as fd:
                json.dump(self.topology, fd)
            fd.close()

    def alley_change(self):
        left_list = self.left_list
        up_table = self.upper_table
        if left_list.selectedItems() and len(left_list.selectedItems()) == 1:
            current_alley = self.topology['alley'][left_list.selectedItems()[0].text()]
            print(left_list.selectedItems()[0].text(), current_alley)

            if sum(eval(up_table.item(2, 1).text())) < int(up_table.item(0, 3).text()):
                QMessageBox.warning(self, "Error", "Не хватает балок!")
            elif sum(eval(up_table.item(2, 1).text())) > int(up_table.item(0, 3).text()):
                QMessageBox.warning(self, "Error", "Не хватает ячеек!")
            else:
                current_alley['rows'] = int(up_table.item(0, 1).text())
                current_alley['columns'] = int(up_table.item(0, 3).text())
                current_alley['start_level'] = int(up_table.item(0, 5).text())
                current_alley['start_cell'] = int(up_table.item(1, 1).text())
                current_alley['count_of_barcodes'] = int(up_table.item(1, 3).text())
                current_alley['local_direction'] = int(up_table.item(1, 5).text())
                current_alley['list_of_balks'] = eval(up_table.item(2, 1).text())
                self.create_table(current_alley['rows'], current_alley['columns'], current_alley['list_of_balks'],
                                  current_alley['start_level'], current_alley['start_cell'],
                                  current_alley['local_direction'])

                with open(self.topology_file, 'w') as fd:
                    json.dump(self.topology, fd)
                fd.close()
        else:
            if not left_list.selectedItems():
                QMessageBox.information(self, "Ошибка!", "Выберите одну аллею из списка!")
            elif left_list.selectedItems() != 1:
                print(len(left_list.selectedItems()))
                QMessageBox.information(self, "Ошибка!", "Выберите ОДНУ аллею из списка!")
            else:
                return

    def photo_checkbox_change(self):
        if self.topology:
            if self.check_box.isChecked() is True:
                self.topology['photo'] = 2
            elif self.check_box.isChecked() is False:
                self.topology['photo'] = 0

            with open(self.topology_file, 'w') as fd:
                json.dump(self.topology, fd)
            fd.close()

    def create_alley_window(self):
        if not self.topology_file:
            QMessageBox.warning(self, "Error", "Не выбран файл с топологией")
        else:
            if not self.new_alley:
                self.new_alley = CreateAlleyWindow(self)
                self.new_alley.show()
            else:
                self.new_alley.show()


class CreateAlleyWindow(QtWidgets.QDialog, Ui_create_alley):
    def __init__(self, parent):
        super(CreateAlleyWindow, self).__init__(parent)
        self.setupUi(self)
        self.settings_dict = None
        self.alley_index = None

        self.alley_buttonbox.rejected.connect(self.close)
        self.alley_buttonbox.accepted.connect(self.accept_button)

    def accept_button(self):
        table = self.new_pallet_table

        for i in range(table.rowCount() - 1):
            if not table.item(i, 1).text():
                QMessageBox.warning(self, "Error", "Некоторые поля пусты")
                return

        self.parse_settings()

        if sum(self.settings_dict['list_of_balks']) != self.settings_dict['columns']:
            QMessageBox.warning(self, "Error", "Ошибка в количестве балок и/или в количестве паллет")
            return
        else:
            self.parent().topology['alley'][self.alley_index] = self.settings_dict
            items = [self.parent().left_list.item(x).text() for x in range(0, self.parent().left_list.count())]
            if self.alley_index not in items:
                self.parent().left_list.itemChanged.disconnect(self.parent().item_changed)

                self.parent().left_list.addItem(self.alley_index)
                n = self.parent().left_list.count()
                self.parent().left_list.item(n - 1).setTextAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                self.parent().left_list.item(n - 1).setFlags(QtCore.Qt.ItemFlag.ItemIsEditable |
                                                             QtCore.Qt.ItemFlag.ItemIsEnabled |
                                                             QtCore.Qt.ItemFlag.ItemIsSelectable)
                self.parent().left_list.itemChanged.connect(self.parent().item_changed)

            with open(self.parent().topology_file, 'w') as fd:
                json.dump(self.parent().topology, fd)
            fd.close()

            self.close()

    def parse_settings(self):
        table = self.new_pallet_table

        self.alley_index = table.item(0, 1).text()
        cols = int(table.item(2, 1).text())
        self.settings_dict = {'rows': int(table.item(1, 1).text()),
                              'columns': cols,
                              'list_of_balks': eval(table.item(3, 1).text()),
                              'start_level': int(table.item(4, 1).text()),
                              'start_cell': int(table.item(5, 1).text()),
                              'count_of_barcodes': int(table.item(6, 1).text()),
                              'local_direction': int(table.item(7, 1).text()),
                              'extra_cells': []}


def main():
    sys.excepthook = my_excepthook
    app = QtWidgets.QApplication(sys.argv)
    window = TopConst()
    window.show()
    app.exec()


def my_excepthook(type_error, value, t_back):
    QMessageBox.critical(TopConst(), "CRITICAL ERROR", str(value))
    sys.__excepthook__(type_error, value, t_back)


if __name__ == '__main__':
    main()
